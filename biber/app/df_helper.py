import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook

@st.cache
def DeSeq2Norm(data):
   ''' DeSeq2 Normalization
   Love et al. (2014) Moderated estimation of fold change and dispersion for RNA-seq data with DESeq2. Genome Biology, 15:550. doi:10.1186/s13059-014-0550-8.
   https://hbctraining.github.io/DGE_workshop/lessons/02_DGE_count_normalization.html
   StatQuest https://www.youtube.com/watch?v=UFB993xufUU
   '''

   #data = pd.read_csv("/Users/daniel/pcanessa/Fast-Review/toyModel.txt",sep=' ')
   #data.set_index('gene', inplace=True)
   # data log
   dataOri = data.copy()
   data = np.log(data)
    
   # average of each row
   #filtered_ave_data = data.copy()
   data['pseudo_reference'] = data.mean(numeric_only=True, axis=1)
   # Filter out all of the genes with -Inf as their average
   # as they will be not used to calculat the median
   data=data[~data['pseudo_reference'].isin([np.nan, np.inf, -np.inf])]

   #Subtract the gene pseudo-references from log counts
   #In this step, you are subtracting the average of the logs from the log of the original data. 
   # Log(x/y) is a ratio.
   ratio_data = data.subtract(data['pseudo_reference'], axis=0)

   # scaling factor = e^median
   scaling_factors = np.exp(ratio_data.mean())

   #Divide the original counts (not log version) by the scaling factors
   dataSeq1 = dataOri.divide(scaling_factors,axis=1)

   return dataSeq1
  

@st.cache(suppress_st_warning=True)
def get_df():
    #file = '/app/data/2022/counts_beb_17032022.csv'
    file = '/app/data/production/counts.csv'
    # get extension and read file
    extension = file.split('.')[1].upper()
    
    expander_ext = st.sidebar.expander(f'{extension} parameters')
    ### TXT TAB ###
    sep ='\t'
    if extension == 'TXT':
        # encoding
        encodings = pd.read_csv('python_encodings.csv')['Encoding'].values
        encoding = expander_ext.selectbox('Select Encoding', encodings, 95)
        # bad lines
        bad_lines_dict = {'Drop':False, 'Break':True}
        i = expander_ext.selectbox('Bad lines', [*bad_lines_dict])
        
        # data frame
        df = pd.read_csv(file, sep=sep, encoding = encoding, error_bad_lines=bad_lines_dict[i])
    ### PICKLE ###
    if extension == 'PICKLE':
        df = pd.read_pickle(file)
    else:
        # Extension parameters expander
        expander_ext = st.sidebar.expander(f'{extension} parameters')
        st.sidebar.text(" \n") #break line
    ### CSV ###
    if extension == 'CSV':
        # encoding
        encodings = pd.read_csv('python_encodings.csv')['Encoding'].values
        encoding = expander_ext.selectbox('Select Encoding', encodings, 95)
        # bad lines
        bad_lines_dict = {'Drop':False, 'Break':True}
        i = expander_ext.selectbox('Bad lines', [*bad_lines_dict])
        # data frame
        df = pd.read_csv(file, encoding = encoding, error_bad_lines=bad_lines_dict[i])
    
    ### EXCEL ###
    elif extension == 'XLSX':
        wb = load_workbook(file)
        # Work sheet
        selected_ws = expander_ext.selectbox('Select Sheet', wb.sheetnames)
        # Start from row
        max_row = wb.worksheets[wb.sheetnames.index(selected_ws)].max_row
        start_r = expander_ext.number_input('Start from row', 1, max_row)
        # data frame
        df = pd.read_excel(file, engine='openpyxl', sheet_name=selected_ws, header=start_r-1)

    df.set_index('target_id', inplace=True)
    df=DeSeq2Norm(df)
    
    return df


def convert_dtypes(df, types, new_types):
    for i, col in enumerate(df.columns):
        # get current column new data type
        new_type = types[new_types[i]]
        if new_type: # check if none
            try: # try to convert
                df[col] = df[col].astype(new_type)
            except: # error msg
                st.write('Could not convert', col, 'to', new_types[i] )
    return df

def handle_nulls(df, null_dict, n_dict):
    for i, col in enumerate(df.columns):
        # check if column is in the null dictionary
        if i in n_dict:
            # get null procedure from dictionary
            # 0 - Remove na
            # 1 - Text
            # 2 - Mean
            # 3 - Median 
            null_proc = null_dict[n_dict[i]]
            try:
                if null_proc == 0:
                    df = df[df[col].notna()]
                elif null_proc == 1:
                    df[col] = df[col].fillna('Not Available')
                elif null_proc == 2:
                    df[col] = df[col].fillna(df[col].mean())
                elif null_proc == 3:
                    df[col] = df[col].fillna(df[col].median())
                elif null_proc == 4:
                    df[col] = df[col].fillna(0)
            except:
                st.write('Could not convert', col, 'nulls to', n_dict[i])
    return df

def handle_duplicates(df, duplicates_dict, action):
    try:
        df = df[~df.duplicated(keep=duplicates_dict[action])]
    except:
        pass
    return df
